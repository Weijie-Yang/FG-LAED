#!/usr/bin/env python
# coding: utf-8



import os
import re
#############cif2input
from utils.cif2input import cif2input
import openpyxl

######## input.data
def split_and_save_files(original_file_path, output_folder):
    original_file_path = os.path.join(original_file_path,'input.data')
    # Read the content of the original file
    with open(original_file_path, 'r') as file:
        content = file.read()

    # Use regular expressions to match the content in comment structure as filenames
    pattern = r'comment structure_(.*?)\n'
    matches = re.findall(pattern, content)

    # Create the target folder if it doesn't exist
    os.makedirs(output_folder, exist_ok=True)

    # Split into blocks
    blocks = content.split('begin\n')[1:]  # # Remove the first element, as the first block before 'begin' is empty

    # Save the split blocks as separate files
    for i, (block, filename) in enumerate(zip(blocks, matches), start=1):
        # Build the output filename, adding the suffix "-in"
        output_filename = os.path.join(output_folder, f'{filename}-in.data')
        with open(output_filename, 'w') as output_file:
            output_file.write('begin\n' + block)

    print(f"{len(blocks)} files created successfully in the '{output_folder}' folder.")



########## config1.data

import shutil
import os

def copy_and_rename_files(destination_folder, source_suffix='-in', destination_suffix='-co'):
    source_file = './template/config1.data'
    # Get all filenames in the destination folder that end with the specified suffix
    in_files = [file for file in os.listdir(destination_folder) if file.endswith(source_suffix + '.data')]

    # Copy files and change the file name suffix
    for in_file in in_files:
        # Build the destination filename
        destination_file = in_file.replace(source_suffix, destination_suffix)
        #  Source and destination file paths
        source_path = source_file
        destination_path = os.path.join(destination_folder, destination_file)
        # Copy the file
        shutil.copy(source_path, destination_path)

    print("Files copied and renamed successfully.")

###########Replace A and B Elements
import os
import re

def replace_values_in_file(file_path, element_values):
    """
    Replace specified values in the file with the electronegativity values of elements
    """
    # Parse the filename to get the chemical symbols for A and B
    file_name = os.path.basename(file_path)
    match = re.match(r'([A-Z][a-z]*)(\d*)([A-Z][a-z]*)(\d*)', file_name)
    if match:
        A_symbol = match.group(1)
        B_symbol = match.group(3)

        # Get the electronegativity values for A and B
        A_electronegativity = element_values.get(A_symbol)
        B_electronegativity = element_values.get(B_symbol)

        if A_electronegativity is not None and B_electronegativity is not None:
            #  Replace values in the file
            with open(file_path, 'r') as file:
                lines = file.readlines()
            with open(file_path, 'w') as file:
                for line in lines:
                    line = re.sub(r'\b0\.00\b', '{:.2f}'.format(A_electronegativity), line)
                    line = re.sub(r'\b9\.00\b', '{:.2f}'.format(B_electronegativity), line)
                    file.write(line)

def process_data_files(folder_path, element_values):
    """
    Process .data files in the specified folder and replace values within them
    """
    for file_name in os.listdir(folder_path):
        # Check if the file ends with .data
        if file_name.endswith('-co.data'):
            file_path = os.path.join(folder_path, file_name)
            replace_values_in_file(file_path, element_values)

# Electronegativity values corresponding to elements
element_values = {
    'H': 2.20, 'He': None,
    'Li': 0.98, 'Be': 1.57, 'B': 2.04, 'C': 2.55, 'N': 3.04, 'O': 3.44, 'F': 3.98, 'Ne': None,
    'Na': 0.93, 'Mg': 1.31, 'Al': 1.61, 'Si': 1.90, 'P': 2.19, 'S': 2.58, 'Cl': 3.16, 'Ar': None,
    'K': 0.82, 'Ca': 1.00, 'Sc': 1.36, 'Ti': 1.54, 'V': 1.63, 'Cr': 1.66, 'Mn': 1.55, 'Fe': 1.83,
    'Co': 1.88, 'Ni': 1.91, 'Cu': 1.90, 'Zn': 1.65, 'Ga': 1.81, 'Ge': 2.01, 'As': 2.18, 'Se': 2.55,
    'Br': 2.96, 'Kr': 3.00, 'Rb': 0.82, 'Sr': 0.95, 'Y': 1.22, 'Zr': 1.33, 'Nb': 1.60, 'Mo': 2.16,
    'Tc': 1.90, 'Ru': 2.20, 'Rh': 2.28, 'Pd': 2.20, 'Ag': 1.93, 'Cd': 1.69, 'In': 1.78, 'Sn': 1.96,
    'Sb': 2.05, 'Te': 2.10, 'I': 2.66, 'Xe': 2.60, 'Cs': 0.79, 'Ba': 0.89, 'La': 1.10, 'Ce': 1.12,
    'Pr': 1.13, 'Nd': 1.14, 'Pm': None, 'Sm': 1.17, 'Eu': None, 'Gd': 1.20, 'Tb': None, 'Dy': 1.22,
    'Ho': 1.23, 'Er': 1.24, 'Tm': 1.25, 'Yb': None, 'Lu': 1.27, 'Hf': 1.30, 'Ta': 1.50, 'W': 2.36,
    'Re': 1.90, 'Os': 2.20, 'Ir': 2.20, 'Pt': 2.28, 'Au': 2.54, 'Hg': 2.00, 'Tl': 1.62, 'Pb': 2.33,
    'Bi': 2.02, 'Po': 2.00, 'At': 2.20, 'Rn': None, 'Fr': 0.70, 'Ra': 0.90, 'Ac': 1.10, 'Th': 1.30,
    'Pa': 1.50, 'U': 1.38, 'Np': 1.36, 'Pu': 1.28, 'Am': 1.30, 'Cm': 1.30, 'Bk': 1.30, 'Cf': 1.30,
    'Es': 1.30, 'Fm': 1.30, 'Md': 1.30, 'No': 1.30, 'Lr': None
}



####################Batch Submission
import os
import sys

import numpy as np
from ase import Atoms
from ase.calculators.singlepoint import SinglePointCalculator
from ase.neighborlist import NeighborList
from utils.ACSF import ACSF

atomic_numbers_dict = {
    1: 2.20, 2: None, 3: 0.98, 4: 1.57, 5: 2.04, 6: 2.55, 7: 3.04, 8: 3.44, 9: 3.98, 10: None, 11: 0.93, 12: 1.31,
    13: 1.61, 14: 1.90, 15: 2.19, 16: 2.58, 17: 3.16, 18: None, 19: 0.82, 20: 1.00, 21: 1.36, 22: 1.54, 23: 1.63,
    24: 1.66, 25: 1.55, 26: 1.83, 27: 1.88, 28: 1.91, 29: 1.90, 30: 1.65, 31: 1.81, 32: 2.01, 33: 2.18, 34: 2.55,
    35: 2.96, 36: 3.00, 37: 0.82, 38: 0.95, 39: 1.22, 40: 1.33, 41: 1.60, 42: 2.16, 43: 1.90, 44: 2.20, 45: 2.28,
    46: 2.20, 47: 1.93, 48: 1.69, 49: 1.78, 50: 1.96, 51: 2.05, 52: 2.10, 53: 2.66, 54: 2.60, 55: 0.79, 56: 0.89,
    57: 1.10, 58: 1.12, 59: 1.13, 60: 1.14, 61: None, 62: 1.17, 63: None, 64: 1.20, 65: None, 66: 1.22, 67: 1.23,
    68: 1.24, 69: 1.25, 70: None, 71: 1.27, 72: 1.3, 73: 1.5, 74: 2.36, 75: 1.9, 76: 2.2, 77: 2.20, 78: 2.28, 79: 2.54,
    80: 2.0, 81: 1.62, 82: 2.33, 83: 2.02, 84: 2.0, 85: 2.2, 86: None, 87: 0.7, 88: 0.9, 89: 1.1, 90: 1.3, 91: 1.5,
    92: 1.38, 93: 1.36, 94: 1.28, 95: 1.3, 96: 1.3, 97: 1.3, 98: 1.3, 99: 1.3, 100: 1.3, 101: 1.3, 102: 1.3, 103: None
}


def read_n2p2(filename='output.data', index=':', with_energy_and_forces='auto'):
    fd = open(filename, 'r')  # @reader decorator ensures this is a file descriptor???
    images = list()
    lineindexlist = []
    lineindex = 0
    line = fd.readline()
    lineindex += 1
    while 'begin' in line:
        lineindexlist.append(lineindex)
        line = fd.readline()
        lineindex += 1
        if 'comment' in line:
            comment = line[7:]
            line = fd.readline()
            lineindex += 1

        cell = np.zeros((3, 3))
        for ii in range(3):
            cell[ii] = [float(jj) for jj in line.split()[1:4]]
            line = fd.readline()
            lineindex += 1

        positions = []
        symbols = []
        charges = []  # not used yet
        nn = []  # not used
        forces = []
        energy = 0.0
        charge = 0.0

        while 'atom' in line:
            sline = line.split()
            positions.append([float(pos) for pos in sline[1:4]])
            symbols.append(sline[4])
            nn.append(float(sline[5]))
            charges.append(float(sline[6]))
            forces.append([float(pos) for pos in sline[7:10]])
            line = fd.readline()
            lineindex += 1

        while 'end' not in line:
            if 'energy' in line:
                energy = float(line.split()[-1])
            if 'charge' in line:
                charge = float(line.split()[-1])
            line = fd.readline()
            lineindex += 1

        image = Atoms(symbols=symbols, positions=positions, cell=cell)

        store_energy_and_forces = False
        if with_energy_and_forces == True:
            store_energy_and_forces = True
        elif with_energy_and_forces == 'auto':
            if energy != 0.0 or np.absolute(forces).sum() > 1e-8:
                store_energy_and_forces = True

        if store_energy_and_forces:
            image.calc = SinglePointCalculator(
                atoms=image,
                energy=energy,
                forces=forces,
                charges=charges)
            # charge  = charge)
        images.append(image)
        # to start the next section
        line = fd.readline()
        lineindex += 1

    if index == ':' or index is None:
        return images, lineindexlist
    else:
        return images[index], lineindexlist


def read_config(configpath):
    cutoff_function = 'tanh'
    gfunctiontypelist = []

    #  Open the file
    with open(configpath, 'r') as file:
        # Read the file content line by line
        for line in file:
            # Remove leading and trailing spaces and newline characters
            line = line.strip()

            # Ignore lines starting with "#"
            if line.startswith("#"):
                continue

            # Add non-empty lines to the list
            if 'symfunction_short' in line:
                gfunctiontypelist.append(line)
            elif 'cutoff_function' in line:
                cutoff_function = line.strip().split()[-1]
            elif 'atom_weighted' in line:
                atom_weighted = int(line.strip().split()[-1])
    return cutoff_function, gfunctiontypelist, atom_weighted


def caculate_gfun(sysatoms, configpath):
    cutoff_function, gfunctiontypelist, atom_weighted = read_config(configpath)
    # print(configpath,':  ',end='')
    # print(atom_weighted)

    atomic_numbers = list(set(sysatoms.get_atomic_numbers()))
    elementlist = sysatoms.get_atomic_numbers()
    gfunction = [[] for _ in range(len(elementlist))]

    for gtype in gfunctiontypelist:
        if gtype.strip().split()[1] == '2':
            atom1, atom2, eta, rs, rc = [float(num_str) for num_str in gtype.strip().split()[2:]]

            electronegativity_list = [atom1]
            ligating_lsit = [atom1, atom2]
            system = []
            for i in atomic_numbers:
                if atomic_numbers_dict[i] in ligating_lsit:
                    system.append(i)
            system1 = []
            for i in atomic_numbers:
                if atomic_numbers_dict[i] in electronegativity_list:
                    system1.append(i)
            symmetry = {'G2': {'eta': [eta, ], 'Rs': [rs]}, }
            bp = ACSF(symmetry, Rc=rc, derivative=True, stress=True, cutoff=cutoff_function,
                      atom_weighted=atom_weighted)
            des = np.sum(bp.calculate(sysatoms, system=system)['x'], axis=1)

            indices = np.where(np.in1d(elementlist, system1))[0]
            for i, k in enumerate(indices):
                gfunction[k].append(des[k])

        elif gtype.strip().split()[1] == '4':
            atom1, atom2, atom3, eta, lambdad, zeta, rc, rs = [float(num_str) for num_str in gtype.strip().split()[2:]]

            electronegativity_list = [atom1]
            ligating_lsit = [atom1, atom2, atom3]
            system = []
            for i in atomic_numbers:
                if atomic_numbers_dict[i] in ligating_lsit:
                    system.append(i)
            system1 = []
            for i in atomic_numbers:
                if atomic_numbers_dict[i] in electronegativity_list:
                    system1.append(i)
            symmetry = {'G4': {'Rs': [rs], 'lambda': [lambdad], 'zeta': [zeta, ], 'eta': [eta, ]}, }
            bp = ACSF(symmetry, Rc=rc, derivative=True, stress=True, cutoff=cutoff_function,
                      atom_weighted=atom_weighted)
            des = np.sum(bp.calculate(sysatoms, system=system)['x'], axis=1)
            indices = np.where(np.in1d(elementlist, system1))[0]
            for i, k in enumerate(indices):
                gfunction[k].append(des[k])
        elif gtype.strip().split()[1] == '5':
            atom1, atom2, atom3, eta, lambdad, zeta, rc, rs = [float(num_str) for num_str in gtype.strip().split()[2:]]

            electronegativity_list = [atom1]
            ligating_lsit = [atom1, atom2, atom3]
            system = []
            for i in atomic_numbers:
                if atomic_numbers_dict[i] in ligating_lsit:
                    system.append(i)
            system1 = []
            for i in atomic_numbers:
                if atomic_numbers_dict[i] in electronegativity_list:
                    system1.append(i)
            symmetry = {'G5': {'Rs': [rs], 'lambda': [lambdad], 'zeta': [zeta, ], 'eta': [eta, ]}, }
            bp = ACSF(symmetry, Rc=rc, derivative=True, stress=True, cutoff=cutoff_function,
                      atom_weighted=atom_weighted)
            des = np.sum(bp.calculate(sysatoms, system=system)['x'], axis=1)
            indices = np.where(np.in1d(elementlist, system1))[0]
            for i, k in enumerate(indices):
                gfunction[k].append(des[k])

    return gfunction


def get_gfuntion(configpath, inputpath, savepath=None):
    # Read structure
    atoms_list, lineindexlist = read_n2p2(filename=inputpath, index=':', with_energy_and_forces='auto')

    if savepath is None:
        savepath = 'output.txt'
    for sysatoms in atoms_list:
        gfunction = caculate_gfun(sysatoms, configpath)

        elementlist = sysatoms.get_atomic_numbers()
        # Open the file for writing
        with open(savepath, 'a') as f:
            # Write the number of non-None elements in gfunction
            f.write(str(len(list(filter(None, gfunction)))) + '\n')
            for ele, sublist in zip(elementlist, gfunction):
                if len(sublist) == 0:
                    continue
                # Format each floating-point number in the list as a string with fixed width
                ele_str = '{:<4}'.format(atomic_numbers_dict[int(ele)])
                line = ele_str + ' ' + ' '.join('{:20.16f}'.format(num) for num in sublist)
                # Write the formatted string to the file
                f.write(line + '\n')

    
def process_files(input_directory):
    output_directory=os.path.join(input_directory, 'out')
    os.makedirs(output_directory, exist_ok=True)
    input_files = [filename for filename in os.listdir(input_directory) if filename.endswith('-in.data')]
    
    for input_file in input_files:
        # Get the base name of the file (excluding suffix and the part after the "-")
        filename_base = input_file.split('-')[0]
        
        # Construct the corresponding configuration file name
        config_file = filename_base + '-co.data'
        
        # Check if the configuration file exists
        if config_file not in os.listdir(input_directory):
            print(f"Warning: Configuration file {config_file} not found for {input_file}. Skipping...")
            continue
        
        inputpath = os.path.join(input_directory, input_file)
        configpath = os.path.join(input_directory, config_file)
        
        # Output file name is the same as the input file name (excluding the suffix)
        output_file = filename_base + '.data'
        savepath = os.path.join(output_directory, output_file)

        get_gfuntion(configpath=configpath, inputpath=inputpath, savepath=savepath)





#########################Identify Sites
import os
import pandas as pd
import re

def find_first_column_of_Uiso(file_path):
    element_dict = {}
    with open(file_path, 'r') as file:
        for line in file:
            if 'Uiso' in line:
                first_column = line.split()[0]
                match = re.match(r'^([A-Za-z]+)', first_column)
                if match:
                    element_name = match.group(1)
                    if element_name not in element_dict:
                        element_dict[element_name] = []
                    element_dict[element_name].append(first_column)
    return element_dict

def process_cif_files(cif_folder_path, sheet_name=0):
    excel_file=os.path.join(cif_folder_path, 'input.xlsx')
    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
    file_idx_dict = {}
    
    for file_name in os.listdir(cif_folder_path):
        if file_name.endswith(".cif"):
            file_path = os.path.join(cif_folder_path, file_name)
            print(f"Processing file: {file_path}")
            
            element_dict = find_first_column_of_Uiso(file_path)
            
            print(f"File: {file_name}")
            for element, columns in element_dict.items():
                matching_row = df[df[0].str.contains(file_name.split('.')[0])]
                if not matching_row.empty:
                    second_column_value = matching_row.iloc[0, 1]
                    for idx, column in enumerate(columns, start=1):
                        matching_elements = [e.strip() for e in second_column_value.split(',') if e.strip() in column]
                        for match in matching_elements:
                            print(f"{idx}. {match}")
                            if file_name not in file_idx_dict:
                                file_idx_dict[file_name] = []
                            file_idx_dict[file_name].append(idx)
                else:
                    print("Element not found in Excel file.")
    return file_idx_dict

def copy_data_files(temp_folder, file_idx_dict):
    data_folder_path=os.path.join(temp_folder, 'out')
    output_folder=d=os.path.join(data_folder_path, 'out1')
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    for file_name, idx_list in file_idx_dict.items():
        data_file_path = os.path.join(data_folder_path, file_name.split('.')[0] + '.data')
        output_file_path = os.path.join(output_folder, file_name.split('.')[0] + '.data')
        
        with open(data_file_path, 'r') as f_in, open(output_file_path, 'w') as f_out:
            for line_num, line in enumerate(f_in, start=1):
                if line_num in [idx + 1 for idx in idx_list]:  
                    f_out.write(line)




    
    
##########################Output as a Table
import pandas as pd
import os

def create_excel_from_data_files(folder_path,folder_path1):
    excel_path=os.path.join(folder_path1, 'out/feature.xlsx')
    all_rows = []
    file_names = []

    for filename in os.listdir(folder_path):
        print('Export to Excel：',filename)
        if filename.endswith('.data'):
            file_names.append(filename)
            file_path = os.path.join(folder_path, filename)

            with open(file_path, 'r') as file:
                content = file.read().strip().split()

            all_rows.append(content)

    df = pd.DataFrame(all_rows)
    
    df.insert(0, 'Formula', file_names)
    df.to_excel(excel_path, index=False)


if __name__ == '__main__':
    #  Specify the path to the source file and the target folder path
    input_file_path = './input'
    temp_folder = './temp/'
    sheet_name = 'top-A'
    # cif2input
    cif2input(input_file_path)
    # Process the input.data file
    split_and_save_files(input_file_path, temp_folder)
    # Process the config.data file
    copy_and_rename_files(temp_folder)
    # Replace element symbols in data files
    process_data_files(temp_folder, element_values)
    # Calculate symmetry functions
    process_files(temp_folder)
    
    #  Identify the site 
    if sheet_name in ['top-A', 'top-B', 'FCC-A','HCP-A','bridge']:
       
        file_idx_dict = process_cif_files(input_file_path)
        copy_data_files(temp_folder, file_idx_dict)
        data_folder_path=os.path.join(temp_folder, 'out/out1')

    else:
        data_folder_path=os.path.join(temp_folder, 'out')
        
    # Export data to Excel
    create_excel_from_data_files(data_folder_path,temp_folder)
####################.Define file paths
cif_folder_path = './input' 
folder_path1 = './temp/'    
file_path = os.path.join(folder_path1, 'out/feature.xlsx')
element_values_path = os.path.join(cif_folder_path, 'input.xlsx')
sheet_name = 'Sheet1'
table_2_file_path = os.path.join(cif_folder_path, 'Feature Table.xlsx')

# Read the element values table from the Excel file
element_values_df = pd.read_excel(element_values_path, sheet_name='top-A')
# Convert the element values table into a dictionary
element_values = element_values_df.set_index('Formula')[['DeltaGOH*', 'DeltaGO*']].to_dict(orient='index')

# Read the content from another Excel file's Sheet1
file_names_df = pd.read_excel(file_path, sheet_name='Sheet1')

# Remove the '.data' part from the 'Formula' column
file_names = file_names_df['Formula'].str.replace('.data', '', regex=True)

# Initialize a new DataFrame to store the new columns
new_columns_df = pd.DataFrame(columns=['Formula', 'DeltaGOH*', 'DeltaGO*'])

# Collect data rows
new_rows = []
for structure_name in file_names:
    # Look up the corresponding values in the element values table
    values = element_values.get(structure_name, None)
    
    if values:
        new_row = {
            'Formula': structure_name,
            'DeltaGOH*': values['DeltaGOH*'],
            'DeltaGO*': values['DeltaGO*']
        }
    else:
        new_row = {
            'Formula': structure_name,
            'DeltaGOH*': None,
            'DeltaGO*': None
        }
    
    new_rows.append(new_row)

#  Add the collected rows to the new DataFrame
new_columns_df = pd.DataFrame(new_rows)

#Replace the original first two columns with the new columns and add the remaining columns
result_df = pd.concat([new_columns_df, file_names_df.iloc[:, 2:]], axis=1)

# Save the updated DataFrame to the same Excel file in Sheet1
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    result_df.to_excel(writer, sheet_name='Sheet1', index=False)


#################### Insert new columns
def insert_columns(file_path, sheet_name):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    new_columns = [
        'A', 'B', 'A-Nm', 'B-Nm', 'A-V', 'B-V', 'Central element', 'Nm',
        'N', 'R (Å)', 'Im (eV)', 'Am (eV)', 'V', 'Sum of Nm of Central and Adjacent Atoms', 
        'Sum of V of Central and Adjacent Atoms'
    ]
    new_data = pd.DataFrame(columns=new_columns)
    df = pd.concat([df.iloc[:, :3], new_data, df.iloc[:, 3:]], axis=1)
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

#################### Extract A, B elements and write to new columns
def extract_and_update_elements(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    elements = []
    max_row = ws.max_row
    for row in range(2, max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value:
            match = re.match(r'([A-Z][a-z]*)(\d*)([A-Z][a-z]*)(\d*)', cell_value)
            if match:
                elements.append((match.group(1), match.group(3)))
    for row, (A_symbol, B_symbol) in enumerate(elements, start=2):
        ws.cell(row=row, column=4).value = A_symbol
        ws.cell(row=row, column=5).value = B_symbol
    wb.save(file_path)
    print("A and B symbols written to columns D and E for all rows from 2 to the last row!")

#################### Update electronegativity and valence
def update_electronegativity_and_valence(file_path, table_2_file_path):
    def read_table_1(file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]
        elements = []
        for row in range(2, ws.max_row + 1):
            element_A = ws.cell(row=row, column=4).value
            element_B = ws.cell(row=row, column=5).value
            elements.append((element_A, element_B))
        wb.close()
        return elements

    def read_table_2(file_path, elements):
        wb = openpyxl.load_workbook(file_path)
        ws = wb['Sheet1']
        element_dict = {ws.cell(row=row, column=1).value: ws.cell(row=row, column=2).value for row in range(3, ws.max_row + 1)}
        column7_dict = {ws.cell(row=row, column=1).value: ws.cell(row=row, column=7).value for row in range(3, ws.max_row + 1)}
        electronegativities = []
        for element_A, element_B in elements:
            elec_A = element_dict.get(element_A, None)
            elec_B = element_dict.get(element_B, None) if element_B else None
            column7_A = column7_dict.get(element_A, None)
            column7_B = column7_dict.get(element_B, None) if element_B else None
            electronegativities.append((elec_A, elec_B, column7_A, column7_B))
        wb.close()
        return electronegativities

    def write_to_table_1(file_path, electronegativities):
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            elec_A, elec_B, col7_A, col7_B = electronegativities[row - 2]
            ws.cell(row=row, column=6).value = elec_A
            ws.cell(row=row, column=7).value = elec_B
            ws.cell(row=row, column=8).value = col7_A
            ws.cell(row=row, column=9).value = col7_B
        wb.save(file_path)
        wb.close()

    elements = read_table_1(file_path)
    if elements:
        electronegativities = read_table_2(table_2_file_path, elements)
        write_to_table_1(file_path, electronegativities)
        print("Electronegativity and valence values written to Table 1 Sheet1 successfully!")
    else:
        print("No data found in Table 1.")

#################### Copy center element column and write to new column
def copy_center_element(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    max_row = ws.max_row
    for row in range(2, max_row + 1):
        value = ws.cell(row=row, column=4).value
        ws.cell(row=row, column=10).value = value
    wb.save(file_path)
    wb.close()
    print("Center elements copied successfully!")

#################### Update center element values
def update_center_element_values(file_path, table_2_file_path):
    def read_table_1(file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]
        elements = []
        for row in range(2, ws.max_row + 1):
            element = ws.cell(row=row, column=10).value
            elements.append(element)
        wb.close()
        return elements

    def read_table_2(file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb['Sheet1']
        properties_dict = {}
        for row in range(3, ws.max_row + 1):
            element = ws.cell(row=row, column=1).value
            if element:
                properties_dict[element] = {
                    'col2': ws.cell(row=row, column=2).value,
                    'col3': ws.cell(row=row, column=3).value,
                    'col4': ws.cell(row=row, column=4).value,
                    'col5': ws.cell(row=row, column=5).value,
                    'col6': ws.cell(row=row, column=6).value,
                    'col7': ws.cell(row=row, column=7).value
                }
        wb.close()
        return properties_dict

    def write_to_table_1(file_path, elements, properties_dict):
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            element = elements[row - 2]
            if element in properties_dict:
                ws.cell(row=row, column=11).value = properties_dict[element]['col2']
                ws.cell(row=row, column=12).value = properties_dict[element]['col3']
                ws.cell(row=row, column=13).value = properties_dict[element]['col4']
                ws.cell(row=row, column=14).value = properties_dict[element]['col5']
                ws.cell(row=row, column=15).value = properties_dict[element]['col6']
                ws.cell(row=row, column=16).value = properties_dict[element]['col7']
        wb.save(file_path)
        wb.close()

    elements = read_table_1(file_path)
    properties_dict = read_table_2(table_2_file_path)
    if elements:
        write_to_table_1(file_path, elements, properties_dict)
        print("Center element values written to Table 1 Sheet1 successfully!")
    else:
        print("No data found in Table 1.")
####################Update coordination electronegativity and valence
def update_excel(file_path, sheet_name):
    # Load workbook
    wb = openpyxl.load_workbook(file_path)
    
    # Select the worksheet
    ws = wb[sheet_name]

    # Get the last row of the sheet
    max_row = ws.max_row

    # Update from row 2 to the last row
    for row in range(2, max_row + 1):
        if (ws.cell(row=row, column=6).value is not None and
            ws.cell(row=row, column=7).value is not None and
            ws.cell(row=row, column=8).value is not None and
            ws.cell(row=row, column=9).value is not None):
            
            # Calculate and write values to columns 17 and 18
            ws.cell(row=row, column=17).value = 3 * ws.cell(row=row, column=7).value + 6 * ws.cell(row=row, column=6).value
            ws.cell(row=row, column=18).value = 3 * ws.cell(row=row, column=9).value + 6 * ws.cell(row=row, column=8).value
    
    # Save workbook
    wb.save(file_path)
    wb.close()
    print("Updated coordination electronegativity and valence in Excel file.")

####################Functions for Processing Columns
def process_columns(file_path, input_sheet_name, output_sheet_name):
    wb = openpyxl.load_workbook(file_path)
    ws_input = wb[input_sheet_name]
    
    if output_sheet_name in wb.sheetnames:
        ws_output = wb[output_sheet_name]
    else:
        ws_output = wb.create_sheet(title=output_sheet_name)
    
    for col in range(1, 19):
        ws_output.cell(row=1, column=col).value = ws_input.cell(row=1, column=col).value
    
    for row in ws_input.iter_rows(min_row=2, max_row=ws_input.max_row, min_col=1, max_col=18):
        for cell in row:
            ws_output.cell(row=cell.row, column=cell.col_idx).value = cell.value
    
    output_col_start = 19
    
    for col in range(19, 83):
        ws_output.cell(row=1, column=output_col_start + (col - 19)).value = f"A*{ws_input.cell(row=1, column=col).value}"
    
    for row in range(2, ws_input.max_row + 1):
        for col in range(19, 83):
            val_17 = ws_input.cell(row=row, column=17).value
            val = ws_input.cell(row=row, column=col).value
            ws_output.cell(row=row, column=output_col_start + (col - 19)).value = val * val_17 if val_17 is not None and val is not None else None

    output_col_start = 19 + (82 - 19 + 1)

    for col in range(19, 83):
        ws_output.cell(row=1, column=output_col_start + (col - 19)).value = f"B*{ws_input.cell(row=1, column=col).value}"
    
    for row in range(2, ws_input.max_row + 1):
        for col in range(19, 83):
            val_18 = ws_input.cell(row=row, column=18).value
            val = ws_input.cell(row=row, column=col).value
            ws_output.cell(row=row, column=output_col_start + (col - 19)).value = val * val_18 if val_18 is not None and val is not None else None
    
    wb.save(file_path)
    wb.close()

####################Delete Unnecessary Columns and Save
def delete_and_save(file_path, input_sheet_name, output_sheet_name):
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    
    # Select the worksheet to process
    ws_input = wb[input_sheet_name]
    
    # Create a new worksheet to save the processed data
    if output_sheet_name in wb.sheetnames:
        wb.remove(wb[output_sheet_name])
    ws_output = wb.create_sheet(title=output_sheet_name)
    
    # Copy data from the input worksheet to the output worksheet
    for row in ws_input.iter_rows():
        values = [cell.value for cell in row]
        ws_output.append(values)
    
    # Copy data from the input worksheet to the output worksheet
    deleted_columns = []  # Used to Store Deleted Column Indices
    for col in range(10, 3, -1):  # Reverse Order Deletion
        ws_output.delete_cols(col)
        deleted_columns.append(col)  # Add Deleted Column Indices
    
    # Delete columns in reverse order 
    for col in range(11, 9, -1):  # Reverse Order Deletion
        ws_output.delete_cols(col)
        deleted_columns.append(col)  # Add Deleted Column Indices
    
    # Save Changes and Close Workbook
    wb.save(file_path)
    wb.close()

    # Print Deleted Column Indices
    print("Deleted columns:", deleted_columns)

# Run All Functions
insert_columns(file_path, sheet_name)
extract_and_update_elements(file_path)
update_electronegativity_and_valence(file_path, table_2_file_path)
copy_center_element(file_path)
update_center_element_values(file_path, table_2_file_path)
update_excel(file_path, sheet_name)

# Process Columns and Save to a New Worksheet
process_columns(file_path, sheet_name, 'Sheet2')

# Delete Unnecessary Columns and Save to a New Worksheet
delete_and_save(file_path, 'Sheet2', 'Sheet3')

